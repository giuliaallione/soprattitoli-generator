"""
main.py – FastAPI. Dopo il parsing chiama Gemini per formattazione e chunking in modo sicuro.
"""

import os, io, uuid, tempfile
from pathlib import Path
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from parser import (parse_document, propose_colors, apply_colors,
                    COLOR_META, split_at_max_chars)
from ai import rework_text, rework_ita_plus

app = FastAPI(title="Soprattitoli Generator API")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_methods=["*"], allow_headers=["*"])

sessions: dict[str, list[dict]] = {}

# ── Modelli ────────────────────────────────────────────────────────────────────

class Criteria(BaseModel):
    session_id: str
    name_caps:   bool = True
    name_bold:   bool = False
    name_custom: str  = ""
    name_sep:    bool = True
    sep_char:    str  = " – "
    desc_italic: bool = True
    desc_parens: bool = True
    max_chars:   int  = 42

class ColorAssignment(BaseModel):
    session_id: str
    assignments: dict[str, str]

class ReworkRequest(BaseModel):
    sentences: list[str]

class UpdateRow(BaseModel):
    session_id: str
    index: int
    field: str
    value: str

class ExportRequest(BaseModel):
    session_id: str

class ItaPlusProcess(BaseModel):
    session_id: str
    column_index: int = 2

class ItaPlusUpdate(BaseModel):
    session_id: str
    row_index: int
    value: str

class ItaPlusSave(BaseModel):
    session_id: str


# ── ITA: Upload ────────────────────────────────────────────────────────────────

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    if not file.filename.endswith(".docx"):
        raise HTTPException(400, "Solo file .docx supportati.")
    content = await file.read()
    sid = str(uuid.uuid4())
    tmp = Path(tempfile.gettempdir()) / "soprattitoli"
    tmp.mkdir(exist_ok=True)
    (tmp / f"{sid}.docx").write_bytes(content)
    return {"session_id": sid, "filename": file.filename}


# ── ITA: Processa + Formattazione AI a Chunk ───────────────────────────────────

@app.post("/api/process")
async def process(criteria: Criteria):
    tmp_path = Path(tempfile.gettempdir()) / "soprattitoli" / f"{criteria.session_id}.docx"
    if not tmp_path.exists():
        raise HTTPException(404, "Sessione non trovata.")
    try:
        rows_from_parser = parse_document(str(tmp_path), {
            "name_caps":   criteria.name_caps,
            "name_bold":   criteria.name_bold,
            "name_custom": criteria.name_custom,
            "name_sep":    criteria.name_sep,
            "sep_char":    criteria.sep_char,
            "desc_italic": criteria.desc_italic,
            "desc_parens": criteria.desc_parens,
        })
    except Exception as e:
        raise HTTPException(500, f"Errore parsing: {e}")

    # Estraiamo le frasi accoppiandole al loro VERO indice originale
    indexed_sentences = [(i, r["ita"]) for i, r in enumerate(rows_from_parser) if r.get("ita") and r["ita"].strip()]
    
    try:
        # L'AI ci restituisce un dizionario: {indice_originale: "testo \n formattato"}
        ai_results_map = rework_text(indexed_sentences)
        
        # Ricostruiamo la lista finale senza sfasare didascalie e battute!
        expanded = []
        for i, row_parser in enumerate(rows_from_parser):
            
            # Formatta il nome in maiuscolo
            personaggio_upper = row_parser.get("personaggio", "").upper()
            
            # Se la riga era una battuta (e AI l'ha elaborata), prendi il testo di AI
            if i in ai_results_map:
                testo_finale = ai_results_map[i]
            else:
                # Altrimenti (es. didascalia) lascia l'originale (che spesso è vuoto o ha solo le note)
                testo_finale = row_parser.get("ita", "")
                
            expanded.append({
                "colore":      row_parser.get("colore", ""),
                "personaggio": personaggio_upper,
                "ita":         testo_finale,
                "note":        row_parser.get("note", "")
            })
            
    except Exception as e:
        print(f"Errore Gemini durante process, uso fallback: {e}")
        expanded = []
        for row in rows_from_parser:
            if row["ita"]:
                for j, line in enumerate(split_at_max_chars(row["ita"], criteria.max_chars)):
                    expanded.append({
                        "colore":      row["colore"],
                        "personaggio": row.get("personaggio", "").upper(),
                        "ita":         line,
                        "note":        row["note"] if j == 0 else "",
                    })
            else:
                expanded.append(row)

    sessions[criteria.session_id] = expanded
    colors = propose_colors(expanded)
    
    return {
        "rows": expanded,
        "total": len(expanded),
        "characters": len({r["personaggio"] for r in expanded if r.get("personaggio")}),
        "proposed_colors": colors,
    }


# ── ITA: Colori ────────────────────────────────────────────────────────────────

@app.post("/api/colors")
async def set_colors(body: ColorAssignment):
    rows = sessions.get(body.session_id)
    if rows is None:
        raise HTTPException(404, "Sessione non trovata.")
    sessions[body.session_id] = apply_colors(rows, body.assignments)
    return {"rows": sessions[body.session_id]}


# ── ITA: Rielabora riga con AI ─────────────────────────────────────────────────

@app.post("/api/rework")
async def rework(body: ReworkRequest):
    if not body.sentences:
        raise HTTPException(400, "Nessuna battuta.")
    try:
        # Crea indici finti per riutilizzare la funzione
        indexed = [(i, s) for i, s in enumerate(body.sentences)]
        res_map = rework_text(indexed)
        # Riporta il dizionario a una lista per questa specifica chiamata
        return {"reworked": [res_map.get(i, orig) for i, orig in enumerate(body.sentences)]}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── ITA: Aggiorna riga ─────────────────────────────────────────────────────────

@app.post("/api/update-row")
async def update_row(body: UpdateRow):
    rows = sessions.get(body.session_id)
    if rows is None:
        raise HTTPException(404, "Sessione non trovata.")
    if body.index < 0 or body.index >= len(rows):
        raise HTTPException(400, "Indice non valido.")
    if body.field not in ("ita", "note", "colore", "personaggio"):
        raise HTTPException(400, "Campo non valido.")
    rows[body.index][body.field] = body.value
    return {"ok": True}


# ── ITA: Esporta ───────────────────────────────────────────────────────────────

@app.post("/api/export")
async def export_excel(body: ExportRequest):
    rows = sessions.get(body.session_id)
    if not rows:
        raise HTTPException(404, "Nessun dato.")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Soprattitoli"
    ws.append(["Colore", "Personaggio", "ITA", "Note"])
    hf = PatternFill(start_color="FF0D47A1", end_color="FF0D47A1", fill_type="solid")
    for cell in ws[1]:
        cell.fill = hf
        cell.font = Font(bold=True, color="FFFFFFFF")
    al = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row.get("colore", ""))
        ws.cell(i, 2, row.get("personaggio", ""))
        ws.cell(i, 3, row.get("ita", ""))
        ws.cell(i, 4, row.get("note", ""))
        ck = row.get("colore", "")
        if ck in COLOR_META:
            hx = COLOR_META[ck]["hex"].lstrip("#")
            ws.cell(i, 1).fill = PatternFill(
                start_color=f"FF{hx}", end_color=f"FF{hx}", fill_type="solid")
        for c in range(1, 5):
            ws.cell(i, c).alignment = al
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 32
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=soprattitoli.xlsx"})


# ── ITA+: Upload ───────────────────────────────────────────────────────────────

@app.post("/api/itaplus/upload")
async def itaplus_upload(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Solo .xlsx supportati.")
    content = await file.read()
    sid = str(uuid.uuid4())
    tmp = Path(tempfile.gettempdir()) / "soprattitoli"
    tmp.mkdir(exist_ok=True)
    (tmp / f"{sid}_itaplus.xlsx").write_bytes(content)
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {"session_id": sid, "filename": file.filename,
            "headers": headers, "rows": ws.max_row - 1}


# ── ITA+: Processa ─────────────────────────────────────────────────────────────

@app.post("/api/itaplus/process")
async def itaplus_process(body: ItaPlusProcess):
    tmp_path = Path(tempfile.gettempdir()) / "soprattitoli" / f"{body.session_id}_itaplus.xlsx"
    if not tmp_path.exists():
        raise HTTPException(404, "File non trovato.")
    wb = openpyxl.load_workbook(str(tmp_path))
    ws = wb.active
    col = body.column_index + 1
    sentences = [str(ws.cell(r, col).value or "") for r in range(2, ws.max_row + 1)]

    non_empty = [(i, s) for i, s in enumerate(sentences) if s.strip()]
    reworked_map = {}
    if non_empty:
        try:
            results = rework_ita_plus([s for _, s in non_empty])
            for idx, (orig_i, _) in enumerate(non_empty):
                reworked_map[orig_i] = results[idx] if idx < len(results) else ""
        except Exception as e:
            raise HTTPException(500, f"Errore Gemini: {e}")

    rows_out = []
    for i, orig in enumerate(sentences):
        rows_out.append({
            "index":    i,
            "original": orig,
            "proposed": reworked_map.get(i, orig),
            "accepted": reworked_map.get(i, orig),
        })

    sessions[f"{body.session_id}_itaplus"] = rows_out
    sessions[f"{body.session_id}_itaplus_col"] = body.column_index
    return {"rows": rows_out, "total": len(rows_out)}


# ── ITA+: Aggiorna proposta ────────────────────────────────────────────────────

@app.post("/api/itaplus/update")
async def itaplus_update(body: ItaPlusUpdate):
    key = f"{body.session_id}_itaplus"
    rows = sessions.get(key)
    if rows is None:
        raise HTTPException(404, "Sessione non trovata.")
    if body.row_index < 0 or body.row_index >= len(rows):
        raise HTTPException(400, "Indice non valido.")
    rows[body.row_index]["accepted"] = body.value
    return {"ok": True}


# ── ITA+: Salva ────────────────────────────────────────────────────────────────

@app.post("/api/itaplus/save")
async def itaplus_save(body: ItaPlusSave):
    key = f"{body.session_id}_itaplus"
    rows = sessions.get(key)
    if rows is None:
        raise HTTPException(404, "Sessione non trovata.")
    tmp_path = Path(tempfile.gettempdir()) / "soprattitoli" / f"{body.session_id}_itaplus.xlsx"
    wb = openpyxl.load_workbook(str(tmp_path))
    ws = wb.active
    new_col = ws.max_column + 1
    ws.cell(1, new_col, "ITA+")
    ws.cell(1, new_col).font = Font(bold=True, color="FFFFFFFF")
    ws.cell(1, new_col).fill = PatternFill(
        start_color="FF0D47A1", end_color="FF0D47A1", fill_type="solid")
    for i, row in enumerate(rows):
        ws.cell(i + 2, new_col, row.get("accepted", ""))
        ws.cell(i + 2, new_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[get_column_letter(new_col)].width = 55
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=soprattitoli_itaplus.xlsx"})


# ── Health ─────────────────────────────────────────────────────────────────────

@app.get("/api/health")
async def health():
    return {"status": "ok"}
